from aiogram import Dispatcher, types
from aiogram.dispatcher import FSMContext
from aiogram.types import InputFile

from tg_bot.DBSM import process_application, add_new_city, all_user_for_rass, namecheck, block, add_service, all_applications_for_review, adminpanel_stat, remont_stat, sell_stat, get_sub_price, change_sub_price
from tg_bot.keyboards import admin_citychoice, tomenu_admin, admin_menu, choice_blok_admin, final_blok_admin, admin_excel_choice, admin_stat_choice, backtostat, adminsub_kb
from tg_bot import generate_random_string
from tg_bot.states import admin

from openpyxl import Workbook
from openpyxl.styles import Font
import pytz, os
from datetime import datetime

def register_admin_handlers(dp: Dispatcher):
    dp.register_callback_query_handler(process_appl, text_startswith = "appl")
    dp.register_message_handler(reason_decline_application, state = admin.appl_reason)
    dp.register_message_handler(service_link, state = admin.service_link)
    dp.register_callback_query_handler(process_mainmenu, text_startswith = "startadmin")
    dp.register_callback_query_handler(process_excel, text_startswith = "excel")
    dp.register_callback_query_handler(process_stat, text_startswith = "stat")
    dp.register_message_handler(process_city, state = admin.city)
    dp.register_message_handler(process_subprice, state = admin.sub_price)
    dp.register_callback_query_handler(citychoice, state = admin.choice_city, text_startswith = "cchoice")
    dp.register_callback_query_handler(proc_blockchoice, state = admin.blok_choice, text_startswith = "block")
    dp.register_message_handler(proc_blockname, state = admin.block_name)
    dp.register_callback_query_handler(proc_blockfinish, state = admin.block_finish, text_startswith = "toblock")
    dp.register_message_handler(proc_blockreason, state = admin.block_reason)
    dp.register_message_handler(procrass, state = admin.rass_text, content_types= types.ContentType.ANY)


async def process_appl(call: types.CallbackQuery, state: FSMContext):
    is_accept = bool(int(call.data.split("_")[1]))
    appl_id = int(call.data.split("_")[2])
    user_id = await process_application(appl_id, is_accept)
    if not is_accept:
        await call.message.answer("Введите причину отклонения заявки 👇")
        await admin.appl_reason.set()
        async with state.proxy() as data:
            data['user_id'] = user_id
    else:
        await call.message.answer("Заявка успешно принята ✅\nСервис, подавший заявку, уже уведомлен об этом 💬")
        await call.message.answer("Введите ссылку на пост в канале с описанием этого сервиса 👇")
        await call.message.bot.send_message(chat_id= user_id, text = "Поздравляем! Ваша заявка была одобрена администратором ✅")
        await admin.service_link.set()
        async with state.proxy() as data:
            data['appl_id'] = appl_id


async def reason_decline_application(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        await message.bot.send_message(chat_id= data['user_id'], text = f"К сожалению, вашу заявку не одобрили ❌\nПричина: <i>{message.text}</i>")
    await message.answer("Заявка успешно отклонена ✅\nСервис, подавший заявку, уже уведомлен о причине 💬")
    await state.finish()


async def service_link(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        await add_service(data['appl_id'], message.text)
    await message.answer("Ссылка закреплена за сервисом")
    await state.finish()





async def process_mainmenu(call: types.CallbackQuery, state: FSMContext):
    action = call.data.split("_")[1]
    if action == "city":
        await call.message.answer("Введите название города")
        await admin.city.set()
    
    elif action == "tomenu":
        await call.message.edit_text("Здравствуйте! Выберите, что вы хотите сделать 👇", reply_markup= admin_menu())

    elif action == "rass":
        await call.message.answer("Отправьте мне сообщение, которое нужно разослать по пользователям 👇")
        await admin.rass_text.set()

    elif action == "blok":
        await call.message.answer("Кого бы вы хотели разблокировать/заблокировать ?", reply_markup= choice_blok_admin())
        await admin.blok_choice.set()
    
    elif action == "appl":
        await call.message.edit_text("Какие заявки вам нужны?", reply_markup= admin_excel_choice())

    elif action == "stat":
        await call.message.edit_text("За какой промежуток времени просмотреть статистику?", reply_markup= admin_stat_choice())

    elif action == "sub":
        price, price_user = await get_sub_price()
        if not price:
            price = "-"
        else:
            price = f"{price.value} ₽"
        if not price_user:
            price_user = "-"
        else:
            price_user = f"{price_user.value} ₽"
        await call.message.edit_text(f"Цена подписки для сервисов - <i>{price}</i>\nЦена подписки для пользователей - <i>{price_user}</i>\n\nХотите изменить цену подписки?", reply_markup= adminsub_kb())
    
    elif action in ("change1", "change0"):
        await admin.sub_price.set()
        async with state.proxy() as data:
            data['is_service'] = int(call.data[-1])
        await call.message.answer("Введите новую цену подписки 👇")



async def process_subprice(message: types.Message, state: FSMContext):
    try:
        if int(message.text) <= 0:
            await message.answer("Введите новую цену подписки натуральным числом без специальных символов")
            return
    except ValueError:
        await message.answer("Введите новую цену подписки натуральным числом без специальных символов")
        return
    
    async with state.proxy() as data:
        is_serv = data['is_service']
        await change_sub_price(is_serv, int(message.text))
    await message.answer(text = "Цена подписки изменена", reply_markup= tomenu_admin())
    await state.finish()


async def process_excel(call: types.CallbackQuery, state: FSMContext):
    action = int(call.data.split("_")[1])
    now_date = datetime.now(pytz.timezone('Europe/Moscow')).strftime('%d.%m.%Y')

    if action: 
        file_name =  await generate_remont_table()
        await call.message.answer_document(document= InputFile(file_name, f"Отчет по заявкам на ремонт {now_date}.xlsx"))
        os.remove(file_name)
        file_name =await generate_sell_table()
        await call.message.answer_document(document= InputFile(file_name, f"Отчет по заявкам на продажу_обмен {now_date}.xlsx"))
        os.remove(file_name)
    else:
        file_name = await generate_applications_table()
        await call.message.answer_document(document= InputFile(file_name, f"Отчет по заявкам на регистрацию {now_date}.xlsx"))
        os.remove(file_name)

async def process_stat(call: types.CallbackQuery, state: FSMContext):
    if call.data == "stat_back":
        await call.message.edit_text("За какой промежуток времени просмотреть статистику?", reply_markup= admin_stat_choice())
        return
        
    method = call.data.split("_")[1]
    methods_js = {"month": "за месяц", "week": "за неделю", "day": "за день", "all": "за все время"}
    res = await adminpanel_stat(method)
    await call.message.edit_text(text = f"Статистика {methods_js[method]}:\n\nКол-во опубликованнх объявлений: {res['count']}\nКол-во завершенных объявлений: {res['closed_count']}\nКол-во объявлений на ремонт: {res['count_remont']}\nКол-во объявлений на продажу/обмен: {res['count_sell']}\nСамый популярный товар(ы): <i>{res['most_common']}</i>", reply_markup= backtostat())


async def process_city(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['city'] = message.text
    await message.answer(f"Вы действительно хотите добавить город {message.text}", reply_markup= admin_citychoice())
    await admin.choice_city.set()


async def citychoice(call: types.CallbackQuery, state: FSMContext):
    is_go  = bool(int(call.data.split("_")[1]))
    if not is_go:
        await state.finish()
        await call.message.answer("Добавление города отменено", reply_markup= tomenu_admin())
    else:
        async with state.proxy() as data:
            await add_new_city(data['city'])
        await call.message.answer("Город добавлен ✅", reply_markup= tomenu_admin())
        await state.finish()


async def procrass(message: types.Message, state: FSMContext):
    count = 0
    for i in await all_user_for_rass():
        try:
            await message.send_copy(chat_id= i)
            count += 1
        except:
            pass

    def get_message_form(n):
        if 11 <= n % 100 <= 19:
            return "сообщений"
        elif n % 10 == 1:
            return "сообщение"
        elif 2 <= n % 10 <= 4:
            return "сообщения"
        else:
            return "сообщений"
        
    await message.answer(f"Разослано {count} {get_message_form(count)}", reply_markup= tomenu_admin())
    await state.finish()


async def proc_blockchoice(call: types.CallbackQuery, state: FSMContext):
    action = int(call.data.split("_")[1])
    if not action:
        await call.message.edit_text("Здравствуйте! Выберите, что вы хотите сделать 👇", reply_markup= admin_menu())
        await state.finish()
        return
    
    async with state.proxy() as data:
        data['is_service_block'] = action == 1
    await admin.block_name.set()
    if action == 1:
        await call.message.answer("Введите юзернейм сервиса в Telegram")
    else:
        await call.message.answer("Введите юзернейм пользователя в Telegram")


async def proc_blockname(message: types.Message, state: FSMContext):
    if message.text == "/start":
        await message.answer("Здравствуйте! Выберите, что вы хотите сделать 👇", reply_markup= admin_menu())
        await state.finish()
        return

    async with state.proxy() as data:
        if not await namecheck(message.text.replace("@", ""), data['is_service_block']):
            if data['is_service_block']:
                await message.answer("К сожалению, сервиса с таким юзернеймом не найдено. Повторите ввод 👇")
            else:
                await message.answer("К сожалению, пользователя с таким юзернеймом не найдено. Повторите ввод 👇")
            return
        data['username'] = message.text.replace("@", "")
    await message.answer("Что вы хотите сделать?", reply_markup= final_blok_admin())
    await admin.block_finish.set()


async def proc_blockfinish(call: types.CallbackQuery, state: FSMContext):
    action = int(call.data.split("_")[1])
    if not action:
        await call.message.edit_text("Здравствуйте! Выберите, что вы хотите сделать 👇", reply_markup= admin_menu())
        await state.finish()
        return

    async with state.proxy() as data:
        user_id = await block(data['username'], data['is_service_block'], action == 1)
        data['block_id'] = user_id
        data['is_block'] = action == 1
        await admin.block_reason.set()
        await call.message.answer("Введите причину блокировки") if action == 1 else await call.message.answer("Введите причину разблокировки")


async def proc_blockreason(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        if data['is_block']:
            await message.bot.send_message(chat_id= data['block_id'], text = f"Вы были заблокированы нашим администратором ❌\nПричина : <i>{message.text}</i>")
            if data['is_service_block']:
                await message.answer(f"Готово! Сервис @{data['username']} заблокирован ❌", reply_markup= tomenu_admin())
            else:
                await message.answer(f"Готово! Пользователь @{data['username']} заблокирован ❌", reply_markup= tomenu_admin())

        else:
            await message.bot.send_message(chat_id= data['block_id'], text = f"Вы были разблокированы нашим администратором ✅\nПричина : <i>{message.text}</i>")
            if data['is_service_block']:
                await message.answer(f"Готово! Сервис @{data['username']} разблокирован ✅", reply_markup= tomenu_admin())
            else:
                await message.answer(f"Готово! Пользователь @{data['username']} разблокирован ✅", reply_markup= tomenu_admin())
    await state.finish()
        


async def generate_applications_table():
    wb = Workbook()
    wb.remove(wb["Sheet"])
    sheet = wb.create_sheet("Заявки на регистрацию", 0)
    sheet["A1"] = "Юзернейм"
    sheet['A1'].font = Font(color="FF0000")  
    sheet["B1"] = "Название"
    sheet['B1'].font = Font(color="FF0000")  
    sheet["C1"] = "Город"
    sheet['C1'].font = Font(color="FF0000")  
    sheet["D1"] = "Адрес"
    sheet['D1'].font = Font(color="FF0000")  
    sheet["E1"] = "Контакты"
    sheet['E1'].font = Font(color="FF0000")  
    sheet["F1"] = "Чем занимается"
    sheet['F1'].font = Font(color="FF0000")  
    sheet["G1"] = "График работы"
    sheet['G1'].font = Font(color="FF0000")  
    sheet["H1"] = "Принята ли заявка"
    sheet['H1'].font = Font(color="FF0000")  
    data = await all_applications_for_review()
    for i in range(len(data)):
        sheet[f"A{i+2}"] = f"@{data[i].username}" if data[i].username is not None else "-"
        sheet[f"B{i+2}"] = data[i].name
        sheet[f"C{i+2}"] = data[i].city
        sheet[f"D{i+2}"] = data[i].adress
        sheet[f"E{i+2}"] = data[i].contacts
        sheet[f"F{i+2}"] = ", ".join(data[i].activity) if len(data[i].activity) != 1 else data[i].activity[0]
        sheet[f"G{i+2}"] = data[i].worktime
        sheet[f"H{i+2}"] = ("✅" if data[i].accepted else "❌") if data[i].accepted is not None else "❓"
    name = "tg_bot/documents/" + generate_random_string(20) + ".xlsx"
    wb.save(name)
    return name 


async def generate_remont_table():
    wb = Workbook()
    wb.remove(wb["Sheet"])
    sheet = wb.create_sheet("Заявки на ремонт", 0)
    sheet["A1"] = "Юзернейм"
    sheet['A1'].font = Font(color="FF0000")  
    sheet["B1"] = "Дата создания"
    sheet['B1'].font = Font(color="FF0000")  
    sheet["C1"] = "Описание"
    sheet['C1'].font = Font(color="FF0000")  
    sheet["D1"] = "Модель"
    sheet['D1'].font = Font(color="FF0000")  
    sheet["E1"] = "Завершена"
    sheet['E1'].font = Font(color="FF0000")  
    sheet["F1"] = "Дата завершения"
    sheet['F1'].font = Font(color="FF0000")  
    sheet["G1"] = "Сервис"
    sheet['G1'].font = Font(color="FF0000")  
    data = await remont_stat()
    for i in range(len(data)):
        sheet[f"A{i+2}"] = f"@{data[i].username}" if data[i].username else "-"
        sheet[f"B{i+2}"] = datetime.strftime(data[i].date_add, '%d.%m.%Y %H:%M')
        sheet[f"C{i+2}"] = data[i].description
        sheet[f"D{i+2}"] = data[i].model
        sheet[f"E{i+2}"] = "✅" if data[i].closed else "❌"
        sheet[f"F{i+2}"] = data[i].date_close if data[i].date_close else "-"
        sheet[f"G{i+2}"] = data[i].closed_with_service if data[i].closed_with_service else "-"
    name = "tg_bot/documents/" + generate_random_string(20) + ".xlsx"
    wb.save(name)
    return name



async def generate_sell_table():
    wb = Workbook()
    wb.remove(wb["Sheet"])
    sheet = wb.create_sheet("Заявки на ремонт", 0)
    sheet["A1"] = "Юзернейм"
    sheet['A1'].font = Font(color="FF0000")  
    sheet["B1"] = "Дата создания"
    sheet['B1'].font = Font(color="FF0000")  
    sheet["C1"] = "Модель"
    sheet["C1"].font = Font(color="FF0000")  
    sheet["D1"] = "Комплект"
    sheet['D1'].font = Font(color="FF0000")  
    sheet["E1"] = "Состояние"
    sheet['E1'].font = Font(color="FF0000")  
    sheet["F1"] = "Аккумулятор"
    sheet['F1'].font = Font(color="FF0000")  
    sheet["G1"] = "Память"
    sheet['G1'].font = Font(color="FF0000")  
    sheet["H1"] = "Размер дисплея"
    sheet['H1'].font = Font(color="FF0000")
    sheet["I1"] = "Завершена"
    sheet['I1'].font = Font(color="FF0000")  
    sheet["J1"] = "Дата завершения"
    sheet['J1'].font = Font(color="FF0000")  
    sheet["K1"] = "Сервис"
    sheet['K1'].font = Font(color="FF0000")  
   
    data = await sell_stat()
    for i in range(len(data)):
        sheet[f"A{i+2}"] = f"@{data[i].username}" if data[i].username else "-"
        sheet[f"B{i+2}"] = datetime.strftime(data[i].date_add, '%d.%m.%Y %H:%M')
        sheet[f"C{i+2}"] = data[i].model
        sheet[f"D{i+2}"] = data[i].equipment
        sheet[f"E{i+2}"] = data[i].condition
        sheet[f"F{i+2}"] = data[i].battery
        sheet[f"G{i+2}"] = data[i].memory
        sheet[f"H{i+2}"] = data[i].display_size
        sheet[f"I{i+2}"] = "✅" if data[i].closed else "❌"
        sheet[f"J{i+2}"] = data[i].date_close if data[i].date_close else "-"
        sheet[f"K{i+2}"] = data[i].closed_with_service if data[i].closed_with_service else "-"
    
    name = "tg_bot/documents/" + generate_random_string(20) + ".xlsx"
    wb.save(name)
    return name

        
        
        



